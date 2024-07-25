#merge excel views.py
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import render, redirect
from .forms import ExcelFileForm
from django.conf import settings
from datetime import datetime

def extract_tables_from_excel_merge(file_paths):
    tables = []

    for file_path in file_paths:
        wb = openpyxl.load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            table = []
            for row in sheet.iter_rows(values_only=True):
                if row[6] is not None:  #  column 6 
                    table.append(list(row))
                else:
                    complete_row = all(cell is not None for cell in row)
                    if complete_row:
                        table.append(list(row))
            if table:  # Add table to tables only if it's not empty
                tables.append(table)

    return tables



def merge_tables(tables):
    merged_table = []
    for i, table in enumerate(tables):
        if i == 0:
            merged_table.extend(table)  # Add the first table as is
        else:
            merged_table.extend(table[1:])  # Skip the header row and add the remaining rows
    return merged_table



def write_merged_table_to_excel(merged_table, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active

    for row_idx, row in enumerate(merged_table, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    wb.save(output_path)

def upload_excel(request):
    if request.method == 'POST':
        files = request.FILES.getlist('excel_files')
        tables = extract_tables_from_excel_merge([file.file for file in files])
        if 'merge_files' in request.POST:
            merged_table = merge_tables(tables)
            document_name = request.POST.get('document_name', 'merged_document')
            output_path = f'Merged/{document_name}_{datetime.now().strftime("%m_%d_%Y_%H%M")}.xlsx'
            write_merged_table_to_excel(merged_table, output_path)
            return render(request, 'word_processor/merge_excel_success.html')
        else:
            # Handle single file upload
            pass
    else:
        form = ExcelFileForm()
    return render(request, 'word_processor/upload_excel.html', {'form': form})
